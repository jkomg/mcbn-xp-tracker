"""Roster → Excel export (roster.export_roster_xlsx)."""

import io

import openpyxl
from flask import Flask, Blueprint
from flask_wtf.csrf import CSRFProtect

import app as app_module
from app.blueprints import roster as roster_module
from app.db import db, DbCharacter, DbLedgerEntry
from app.db_service import DBService

_fake_dashboard_bp = Blueprint('dashboard', __name__)


@_fake_dashboard_bp.route('/login')
def login():
    return 'login', 200


def _app():
    app = Flask(__name__)
    app.config['TESTING'] = True
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///:memory:'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['WTF_CSRF_ENABLED'] = False
    app.secret_key = 'test-secret'
    db.init_app(app)
    CSRFProtect().init_app(app)
    service = DBService()
    app_module.db_service = service
    roster_module.db_service = service
    roster_module.sheets_sync = None
    app.register_blueprint(roster_module.bp, url_prefix='/roster')
    app.register_blueprint(_fake_dashboard_bp)
    with app.app_context():
        db.create_all()
    return app


def _seed(app):
    with app.app_context():
        db.session.add(DbCharacter(
            character_name='Tulip Miller', player_discord_name='Alice', clan='Tremere',
            age_category='Neonate', sect='Camarilla', active=True, status='active',
            creation_xp=10, enemy='', date_added='20260101', notes='',
        ))
        db.session.add(DbCharacter(
            character_name='Bram Stoker', player_discord_name='Bob', clan='Nosferatu',
            age_category='Ancilla', sect='Anarch', active=False, status='retired',
            creation_xp=15, enemy='', date_added='20260102', notes='',
        ))
        db.session.add(DbLedgerEntry(
            character_name='Tulip Miller', date='20260105', awarded=5, spent=0,
            reason='Session XP', entered_by='ST', timestamp='20260105 12:00:00',
        ))
        db.session.commit()


def _staff_client(app):
    client = app.test_client()
    with client.session_transaction() as sess:
        sess['authenticated'] = True
        sess['staff_user'] = 'Tester'
    return client


def _load_workbook(resp):
    return openpyxl.load_workbook(io.BytesIO(resp.data))


def test_export_default_active_filter_excludes_inactive():
    app = _app()
    _seed(app)
    client = _staff_client(app)

    resp = client.get('/roster/export.xlsx')
    assert resp.status_code == 200
    assert resp.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    wb = _load_workbook(resp)
    ws = wb.active
    names = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value]
    assert names == ['Tulip Miller']


def test_export_all_includes_inactive_and_xp_columns():
    app = _app()
    _seed(app)
    client = _staff_client(app)

    resp = client.get('/roster/export.xlsx?show=all')
    wb = _load_workbook(resp)
    ws = wb.active

    header = [c.value for c in ws[1]]
    assert header == [
        'Character Name', 'Player', 'Clan', 'Age Category', 'Sect', 'Status',
        'Creation XP', 'Earned XP', 'Approved Spends', 'Available XP',
        'Portrait', 'Enemy', 'Date Added', 'Notes',
    ]

    rows = {row[0].value: row for row in ws.iter_rows(min_row=2) if row[0].value}
    assert set(rows) == {'Tulip Miller', 'Bram Stoker'}

    tulip = rows['Tulip Miller']
    assert tulip[1].value == 'Alice'
    assert tulip[5].value == 'Active'
    assert tulip[6].value == 10  # creation xp
    assert tulip[7].value == 5   # earned xp from ledger
    assert tulip[9].value == 15  # available xp = 10 + 5

    bram = rows['Bram Stoker']
    assert bram[5].value == 'Retired'


def test_export_respects_clan_filter():
    app = _app()
    _seed(app)
    client = _staff_client(app)

    resp = client.get('/roster/export.xlsx?show=all&clan=Nosferatu')
    wb = _load_workbook(resp)
    ws = wb.active
    names = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value]
    assert names == ['Bram Stoker']


def test_export_neutralizes_formula_injection_in_free_text_fields():
    """A player-supplied name/player/enemy/notes value starting with "="
    must not open the workbook as an evaluated formula — openpyxl
    auto-flags leading "=" strings as formulas unless forced back to
    plain text (see export_roster_xlsx)."""
    app = _app()
    with app.app_context():
        db.session.add(DbCharacter(
            character_name='=1+1', player_discord_name='=cmd|calc',
            clan='', age_category='', sect='', active=True, status='active',
            creation_xp=0, enemy='=SUM(A1:A9)', date_added='', notes='=HYPERLINK("evil")',
        ))
        db.session.commit()
    client = _staff_client(app)

    resp = client.get('/roster/export.xlsx?show=all')
    wb = _load_workbook(resp)
    ws = wb.active
    row = next(r for r in ws.iter_rows(min_row=2) if r[0].value == '=1+1')

    assert row[0].data_type == 's'
    assert row[1].value == '=cmd|calc' and row[1].data_type == 's'
    assert row[11].value == '=SUM(A1:A9)' and row[11].data_type == 's'
    assert row[13].value == '=HYPERLINK("evil")' and row[13].data_type == 's'
